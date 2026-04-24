from rest_framework.viewsets import ViewSet, ModelViewSet
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework import status
from django.utils import timezone
from django.db.models import Sum
from datetime import timedelta
from decimal import Decimal

from orders.models import Order
from users.models import User
from branch.models import Branch
from .models import DriverSettlement, SettlementDetail
from .serializers import DriverSettlementSerializer, DriverSettlementListSerializer


class SettlementViewSet(ViewSet):
    """
    ViewSet untuk generate dan menampilkan settlement reports.
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    
    @action(detail=False, methods=['get'])
    def generate_settlement(self, request):
        """
        Generate settlement untuk semua driver dalam periode tertentu.
        
        Query params:
        - driver_id: filter untuk driver tertentu
        - branch_id: filter untuk branch tertentu
        - period: '5_days', '15_days', '30_days'
        - start_date: tanggal mulai (YYYY-MM-DD)
        - end_date: tanggal akhir (YYYY-MM-DD)
        """
        driver_id = request.query_params.get('driver_id')
        branch_id = request.query_params.get('branch_id')
        period = request.query_params.get('period') or '30_days'
        
        # Tentukan date range
        if period == '5_days':
            start_date = timezone.now().date() - timedelta(days=5)
        elif period == '15_days':
            start_date = timezone.now().date() - timedelta(days=15)
        else:  # 30_days
            start_date = timezone.now().date() - timedelta(days=30)
        
        end_date = timezone.now().date()
        
        # Get drivers (filter jika ada)
        drivers = User.objects.filter(role='driver')
        if driver_id:
            drivers = drivers.filter(id=driver_id)
        
        # Get branches (filter jika ada)
        branches = Branch.objects.all()
        if branch_id:
            branches = branches.filter(id=branch_id)
        
        rekap = []
        
        for driver in drivers:
            for branch in branches:
                # Get orders untuk driver ini dalam periode ini dan branch ini
                orders = Order.objects.filter(
                    driver=driver,
                    branch=branch,
                    status__in=['done', 'on_delivery'],  # Hanya order yang selesai
                    created_at__date__gte=start_date,
                    created_at__date__lte=end_date
                )
                
                if not orders.exists():
                    continue
                
                # Cek jika settlement sudah ada
                settlement, created = DriverSettlement.objects.get_or_create(
                    driver=driver,
                    branch=branch,
                    period=period,
                    period_start=start_date,
                    period_end=end_date
                )
                
                # Hitung settlement
                settlement.calculate_settlement(orders)
                settlement.save()
                
                # Buat atau update detail items
                SettlementDetail.objects.filter(settlement=settlement).delete()
                
                for order in orders:
                    # Estimasi tarif dan service fee
                    estimated_tarif = order.final_price * Decimal('0.8') if order.final_price else Decimal(0)
                    estimated_service_fee = order.final_price * Decimal('0.2') if order.final_price else Decimal(0)
                    
                    deduction = settlement._calculate_order_deduction(estimated_tarif)
                    
                    SettlementDetail.objects.create(
                        settlement=settlement,
                        order=order,
                        order_code=order.order_code,
                        tarif=estimated_tarif,
                        service_fee=estimated_service_fee,
                        total_price=order.final_price or Decimal(0),
                        deduction=deduction,
                        settlement_amount=(order.final_price or Decimal(0)) - deduction,
                        pickup_location=order.pickup_location,
                        drop_location=order.drop_location,
                        completed_at=order.created_at
                    )
                
                rekap.append(settlement)
        
        serializer = DriverSettlementListSerializer(rekap, many=True)
        return Response({
            'status': 'generated',
            'count': len(rekap),
            'period': period,
            'date_range': {
                'start': start_date,
                'end': end_date
            },
            'settlements': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def list_settlements(self, request):
        """
        List semua rekap dengan filter dan pagination.
        """
        driver_id = request.query_params.get('driver_id')
        branch_id = request.query_params.get('branch_id')
        period = request.query_params.get('period')
        status_filter = request.query_params.get('status')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        rekap = DriverSettlement.objects.select_related('driver', 'branch')
        
        if driver_id:
            rekap = rekap.filter(driver_id=driver_id)
        if branch_id:
            rekap = rekap.filter(branch_id=branch_id)
        if period:
            rekap = rekap.filter(period=period)
        if status_filter:
            rekap = rekap.filter(status=status_filter)
        if start_date:
            rekap = rekap.filter(period_start__gte=start_date)
        if end_date:
            rekap = rekap.filter(period_end__lte=end_date)
        
        serializer = DriverSettlementListSerializer(rekap, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def settlement_detail(self, request):
        """
        Detail settlement untuk satu driver dan periode.
        """
        settlement_id = request.query_params.get('settlement_id')
        
        if not settlement_id:
            return Response({'error': 'settlement_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            settlement = DriverSettlement.objects.get(id=settlement_id)
        except DriverSettlement.DoesNotExist:
            return Response({'error': 'Settlement not found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = DriverSettlementSerializer(settlement)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def mark_paid(self, request):
        """
        Mark settlement as paid.
        """
        settlement_id = request.data.get('settlement_id')
        
        if not settlement_id:
            return Response({'error': 'settlement_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            settlement = DriverSettlement.objects.get(id=settlement_id)
            settlement.status = 'paid'
            settlement.paid_date = timezone.now().date()
            settlement.save()
            
            serializer = DriverSettlementSerializer(settlement)
            return Response(serializer.data)
        except DriverSettlement.DoesNotExist:
            return Response({'error': 'Settlement not found'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Export settlement detail ke Excel format.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        driver_id = request.query_params.get('driver_id')
        branch_id = request.query_params.get('branch_id')
        period = request.query_params.get('period')
        status_filter = request.query_params.get('status')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        details = SettlementDetail.objects.select_related('settlement', 'settlement__driver', 'settlement__branch', 'order')
        
        if driver_id:
            details = details.filter(settlement__driver_id=driver_id)
        if branch_id:
            details = details.filter(settlement__branch_id=branch_id)
        if period:
            details = details.filter(settlement__period=period)
        if status_filter:
            details = details.filter(settlement__status=status_filter)
        if start_date:
            details = details.filter(settlement__period_start__gte=start_date)
        if end_date:
            details = details.filter(settlement__period_end__lte=end_date)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Settlement Detail'
        
        # Header
        headers = [
            'Tanggal', 'Order ID', 'Driver', 'Branch', 'Jarak (km)',
            'Tarif Dasar', 'Service', 'Total', 'Potongan', 'Net Driver',
            'Status', 'Period', 'Date Range'
        ]
        
        header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for idx, detail in enumerate(details, 2):
            settlement = detail.settlement
            ws.cell(row=idx, column=1).value = detail.completed_at.date() if detail.completed_at else settlement.period_start
            ws.cell(row=idx, column=2).value = detail.order_code
            ws.cell(row=idx, column=3).value = settlement.driver.get_full_name() or settlement.driver.username
            ws.cell(row=idx, column=4).value = settlement.branch.name
            ws.cell(row=idx, column=5).value = float(detail.distance)
            ws.cell(row=idx, column=6).value = float(detail.tarif)
            ws.cell(row=idx, column=7).value = float(detail.service_fee)
            ws.cell(row=idx, column=8).value = float(detail.total_price)
            ws.cell(row=idx, column=9).value = float(detail.deduction)
            ws.cell(row=idx, column=10).value = float(detail.settlement_amount)
            ws.cell(row=idx, column=11).value = settlement.get_status_display()
            ws.cell(row=idx, column=12).value = settlement.get_period_display()
            ws.cell(row=idx, column=13).value = f"{settlement.period_start} - {settlement.period_end}"

            for col in [5, 6, 7, 8, 9, 10]:
                ws.cell(row=idx, column=col).number_format = '#,##0.00' if col == 5 else '#,##0'

            if float(detail.deduction) > 10000:
                ws.cell(row=idx, column=9).font = Font(color='FF0000', bold=True)
            if float(detail.total_price) >= 100000:
                ws.cell(row=idx, column=10).font = Font(color='008000', bold=True)

        # Footer totals
        total_row = details.count() + 2
        ws.cell(row=total_row, column=7).value = 'TOTAL:'
        ws.cell(row=total_row, column=7).font = Font(bold=True)
        ws.cell(row=total_row, column=8).value = float(details.aggregate(total=Sum('total_price'))['total'] or 0)
        ws.cell(row=total_row, column=9).value = float(details.aggregate(deduction=Sum('deduction'))['deduction'] or 0)
        ws.cell(row=total_row, column=10).value = float(details.aggregate(net=Sum('settlement_amount'))['net'] or 0)
        for col in [8, 9, 10]:
            ws.cell(row=total_row, column=col).number_format = '#,##0'
            ws.cell(row=total_row, column=col).font = Font(bold=True)

        # Adjust column widths
        widths = [14, 18, 22, 18, 12, 15, 15, 15, 15, 15, 14, 14, 22]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        filename = 'settlement_detail'
        if period:
            filename += f'_{period}'
        if status_filter:
            filename += f'_{status_filter}'
        filename += f'_{timezone.now().strftime("%Y%m%d")}.xlsx'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        wb.save(response)
        return response

        
        # Response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=settlements.xlsx'
        wb.save(response)
        
        return response
